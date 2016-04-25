from new_tasks import report_hotmail
from openpyxl import Workbook
import requests, datetime
from celery import group
import time
from test import spf_check_task
from celery.result import AsyncResult


def get_planning(date=datetime.date.today() + datetime.timedelta(days=2)):
    wb = Workbook()
    ws = wb.active
    day_name = date.strftime("%A")
    url = "http://192.168.0.250/workers/ajax.php"
    shifts = {"15": "Day Off", "16": "09H", "17": "14H30", "21": "Congé", "25": "05H", "26": "Maladie"}
    shift_colors = {"15": "999999", "16": "66ff66", "17": "ff9900", "21": "ff6666", "25": "33ccff", "26": "ffff00"}
    teams = requests.post(url, data={"what": "getteams"}).json()

    ws['D2'] = date

    ws['B4'] = "Nom & Prénom"
    ws['C4'] = day_name

    i = 5
    for team in teams:
        print(("*" * 25 + "%s" + "*" * 25) % team.get("team_name", None))
        members = requests.post(url, data={"what": "getmembers", "team": team.get("team", None)}).json()
        for member in members:
            member_name = member.get("name", None)
            shift = shifts.get(
                str(requests.post(url, data={"what": "getshft", "id": member.get("id", None), "date": date}).json()),
                None)
            ws['B%s' % str(i)] = member_name
            ws['C%s' % str(i)] = shift
            print("%s: %s" % (member_name, shift))
            i += 1
    wb.save("%s_%s.xlsx" % (day_name, str(date)))


# get_planning()

report_hotmail(actions='SS', subject='notifica',
               email={'login': 'ellansourwinebg2194@hotmail.com', 'password': 'RUbg3491'}, proxy=None)
# report_hotmail(actions='FM,CLS', subject='a', email={'login': 'jonathan13247@hotmail.com', 'password': 'cvc22016'}, proxy=None)

# tas = group(report_hotmail.s(actions='', subject='Anniversaire', proxy=None,
#                              email={'login': 'juliannadauermy5880@hotmail.com', 'password': 'WEmy5612'})
#             .set(queue='smghanen') for i in range(1))
#
# tas_results = tas.apply_async()
# time.sleep(60)
# for tas_result in tas_results:
#     print(tas_result.state)





i = 1
with open("1m-ANAS.txt", 'r') as f:
    for domain in f:
        try:
            spf_task = spf_check_task.apply_async((str(i), domain.replace("\n", "")), queue="SPF63")
            task_file_id = open("1m-ANAS_id.txt", 'a')
            task_file_id.write("%s;%s\n" % (str(i), spf_task.id))
            task_file_id.close()
            i += 1
        except Exception as ex:
            continue

i = 1
with open("1m-2-ANAS.txt", 'r') as f:
    for domain in f:
        try:
            spf_task = spf_check_task.apply_async((str(i), domain.replace("\n", "")), queue="SPF63")
            task_file_id = open("1m-2-ANAS_id.txt", 'a')
            task_file_id.write("%s;%s\n" % (str(i), spf_task.id))
            task_file_id.close()
            i += 1
        except Exception as ex:
            continue

i = 1
with open("1M-bazi.txt", 'r') as f:
    for domain in f:
        try:
            spf_task = spf_check_task.apply_async((str(i), domain.replace("\n", "")), queue="SPF63")
            task_file_id = open("1M-bazi_id.txt", 'a')
            task_file_id.write("%s;%s\n" % (str(i), spf_task.id))
            task_file_id.close()
            i += 1
        except Exception as ex:
            continue

i = 1
with open("1m-ANAS_id.txt", 'r') as f:
    for idd in f:
        spf = AsyncResult(idd.replace("\n", "").split(";")[1])
        if spf.state == 'SUCCESS':
            try:
                task_file_results = open("SPF_1m-ANAS.txt", 'a')
                task_file_results.write("%s\n" % spf.result)
                task_file_results.close()
            except UnicodeEncodeError:
                for character in spf.result:
                    try:
                        task_file_results = open("SPF_1m-ANAS.txt", 'a')
                        task_file_results.write(character)
                        task_file_results.close()
                    except UnicodeEncodeError:
                        continue
                task_file_results = open("SPF_1m-ANAS.txt", 'a')
                task_file_results.write("\n")
                task_file_results.close()
        else:
            task_file_results = open("SPF_ERRORS_1m-ANAS.txt", 'a')
            task_file_results.write(idd)
            task_file_results.close()

i = 1
with open("1m-2-ANAS_id.txt", 'r') as f:
    for idd in f:
        spf = AsyncResult(idd.replace("\n", "").split(";")[1])
        if spf.state == 'SUCCESS':
            try:
                task_file_results = open("SPF_1m-2-ANAS.txt", 'a')
                task_file_results.write("%s\n" % spf.result)
                task_file_results.close()
            except UnicodeEncodeError:
                for character in spf.result:
                    try:
                        task_file_results = open("SPF_1m-2-ANAS.txt", 'a')
                        task_file_results.write(character)
                        task_file_results.close()
                    except UnicodeEncodeError:
                        continue
                task_file_results = open("SPF_1m-2-ANAS.txt", 'a')
                task_file_results.write("\n")
                task_file_results.close()
        else:
            task_file_results = open("SPF_ERRORS_1m-2-ANAS.txt", 'a')
            task_file_results.write(idd)
            task_file_results.close()

i = 1
with open("1M-bazi_id.txt", 'r') as f:
    for idd in f:
        spf = AsyncResult(idd.replace("\n", "").split(";")[1])
        if spf.state == 'SUCCESS':
            try:
                task_file_results = open("SPF_1M-bazi.txt", 'a')
                task_file_results.write("%s\n" % spf.result)
                task_file_results.close()
            except UnicodeEncodeError:
                for character in spf.result:
                    try:
                        task_file_results = open("SPF_1M-bazi.txt", 'a')
                        task_file_results.write(character)
                        task_file_results.close()
                    except UnicodeEncodeError:
                        continue
                task_file_results = open("SPF_1M-bazi.txt", 'a')
                task_file_results.write("\n")
                task_file_results.close()
        else:
            task_file_results = open("SPF_ERRORS_1M-bazi.txt", 'a')
            task_file_results.write(idd)
            task_file_results.close()


